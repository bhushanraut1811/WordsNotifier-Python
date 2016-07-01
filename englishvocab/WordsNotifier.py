from random import random, randint
from time import sleep
import openpyxl
from PyQt4 import Qt
import sys

'''
--LEARN--
# print wb.get_sheet_names()
# print sheet.title
'''

wb = openpyxl.load_workbook('EnglishVocabList.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

rowCount = sheet.max_row
wordList = []
word = ''
app = Qt.QApplication(sys.argv)


def parseWordsExcel():
    for index in xrange(1, rowCount):
        word = sheet.cell(row=index, column=1).value + ' : ' + sheet.cell(row=index, column=2).value
        # print word
        wordList.append(word)


def showWordPopUp():
    print 'hello'
    systemtray_icon = Qt.QSystemTrayIcon(Qt.QIcon('word_icon.png'))
    systemtray_icon.show()
    systemtray_icon.showMessage('Learn this Word', wordList[randint(0, rowCount)])
    sleep(3601)


parseWordsExcel()
while True:
    showWordPopUp()
